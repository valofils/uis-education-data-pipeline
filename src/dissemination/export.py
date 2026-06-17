"""
export.py
Export the imputed analytical dataset in CSV, JSON, and Excel formats.

Excel workbook sheets
---------------------
1. Data        : full imputed wide table
2. Coverage    : indicator coverage statistics
3. Imputation  : imputation transparency report
4. Metadata    : indicator concept descriptions

Location: src/dissemination/export.py
"""

import json
import logging
import pandas as pd
from pathlib import Path

from src.utils.config import (
    DATA_PROCESSED_DIR,
    OUTPUTS_EXPORTS_DIR,
    OUTPUTS_REPORTS_DIR,
    METADATA_FILE,
    SDG4_INDICATORS,
    EXPORT_FORMATS,
    LOG_FORMAT,
    LOG_DATE_FMT,
)

# ── Logger ────────────────────────────────────────────────────────────────────
logging.basicConfig(format=LOG_FORMAT, datefmt=LOG_DATE_FMT, level=logging.INFO)
log = logging.getLogger(__name__)

# ── Input files ───────────────────────────────────────────────────────────────
IMPUTED_FILE   = DATA_PROCESSED_DIR  / "imputed_wide.csv"
COVERAGE_FILE  = OUTPUTS_REPORTS_DIR / "coverage_stats.csv"  
IMPUTATION_RPT = OUTPUTS_REPORTS_DIR / "imputation_report.csv"

# ── Output files ──────────────────────────────────────────────────────────────
OUT_CSV   = OUTPUTS_EXPORTS_DIR / "uis_sdg4_indicators.csv"
OUT_JSON  = OUTPUTS_EXPORTS_DIR / "uis_sdg4_indicators.json"
OUT_EXCEL = OUTPUTS_EXPORTS_DIR / "uis_sdg4_indicators.xlsx"


# ── Loaders ───────────────────────────────────────────────────────────────────
def load_data() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Load imputed data, coverage stats, and imputation report."""
    if not IMPUTED_FILE.exists():
        log.info("Imputed file not found — running imputation step.")
        from src.imputation.impute import run as impute_run
        impute_run()

    df       = pd.read_csv(IMPUTED_FILE, dtype={"YEAR": "Int64"})
    coverage = pd.read_csv(COVERAGE_FILE) if COVERAGE_FILE.exists() else pd.DataFrame()
    imp_rpt  = pd.read_csv(IMPUTATION_RPT) if IMPUTATION_RPT.exists() else pd.DataFrame()

    log.info("Loaded imputed data: %s rows × %s columns.", *df.shape)
    return df, coverage, imp_rpt


def load_metadata_sheet() -> pd.DataFrame:
    """Extract indicator concept sheet from metadata registry JSON."""
    if not METADATA_FILE.exists():
        log.warning("Metadata registry not found — skipping metadata sheet.")
        return pd.DataFrame()

    with open(METADATA_FILE, "r", encoding="utf-8") as f:
        registry = json.load(f)

    concepts = registry.get("concept_scheme", {}).get("concepts", {})
    rows = []
    for ind_id, meta in concepts.items():
        rows.append({
            "INDICATOR_ID":       ind_id,
            "LABEL":              meta.get("label", ""),
            "SDG_TARGET":         meta.get("sdg_target", ""),
            "UNIT":               meta.get("unit", ""),
            "MEASUREMENT":        meta.get("measurement", ""),
            "DISAGGREGATIONS":    ", ".join(meta.get("disaggregations", [])),
            "SOURCE_ORGANISATION":meta.get("source_organisation", ""),
            "NOTE":               meta.get("note", ""),
        })

    return pd.DataFrame(rows)


# ── Export functions ──────────────────────────────────────────────────────────
def export_csv(df: pd.DataFrame) -> Path:
    """Export to CSV."""
    df.to_csv(OUT_CSV, index=False)
    log.info("CSV  exported: %s  (%s KB)", OUT_CSV.name,
             round(OUT_CSV.stat().st_size / 1024, 1))
    return OUT_CSV


def export_json(df: pd.DataFrame) -> Path:
    """
    Export to JSON — records orientation with metadata wrapper.
    """
    payload = {
        "dataset":      "UIS SDG 4 Education Indicators",
        "generated":    pd.Timestamp.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
        "n_records":    len(df),
        "indicators":   SDG4_INDICATORS,
        "records":      json.loads(
            df.to_json(orient="records", date_format="iso")
        ),
    }

    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False)

    log.info("JSON exported: %s  (%s KB)", OUT_JSON.name,
             round(OUT_JSON.stat().st_size / 1024, 1))
    return OUT_JSON


def export_excel(
    df: pd.DataFrame,
    coverage: pd.DataFrame,
    imp_rpt: pd.DataFrame,
    meta_sheet: pd.DataFrame,
) -> Path:
    """
    Export to Excel workbook with four sheets and basic formatting.
    """
    with pd.ExcelWriter(OUT_EXCEL, engine="openpyxl") as writer:

        # ── Sheet 1: Data ─────────────────────────────────────────────────────
        df.to_excel(writer, sheet_name="Data", index=False)
        ws = writer.sheets["Data"]
        _format_header(ws)
        _autofit_columns(ws)

        # ── Sheet 2: Coverage ─────────────────────────────────────────────────
        if not coverage.empty:
            coverage.to_excel(writer, sheet_name="Coverage", index=False)
            ws2 = writer.sheets["Coverage"]
            _format_header(ws2)
            _autofit_columns(ws2)

        # ── Sheet 3: Imputation ───────────────────────────────────────────────
        if not imp_rpt.empty:
            imp_rpt.to_excel(writer, sheet_name="Imputation", index=False)
            ws3 = writer.sheets["Imputation"]
            _format_header(ws3)
            _autofit_columns(ws3)

        # ── Sheet 4: Metadata ─────────────────────────────────────────────────
        if not meta_sheet.empty:
            meta_sheet.to_excel(writer, sheet_name="Metadata", index=False)
            ws4 = writer.sheets["Metadata"]
            _format_header(ws4)
            _autofit_columns(ws4)

    log.info("Excel exported: %s  (%s KB)", OUT_EXCEL.name,
             round(OUT_EXCEL.stat().st_size / 1024, 1))
    return OUT_EXCEL


# ── Excel formatting helpers ──────────────────────────────────────────────────
def _format_header(ws) -> None:
    """Bold and colour the header row."""
    from openpyxl.styles import Font, PatternFill, Alignment

    header_fill = PatternFill(
        start_color="1F4E79", end_color="1F4E79", fill_type="solid"
    )
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for cell in ws[1]:
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[1].height = 20


def _autofit_columns(ws, max_width: int = 30) -> None:
    """Set column widths based on content length."""
    for col in ws.columns:
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=8,
        )
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, max_width)


# ── Main entry ────────────────────────────────────────────────────────────────
def run() -> dict[str, Path]:
    """
    Export imputed data in all configured formats.

    Returns
    -------
    Dict mapping format name to output Path.
    """
    df, coverage, imp_rpt = load_data()
    meta_sheet = load_metadata_sheet()

    outputs = {}

    if "csv" in EXPORT_FORMATS:
        outputs["csv"] = export_csv(df)

    if "json" in EXPORT_FORMATS:
        outputs["json"] = export_json(df)

    if "xlsx" in EXPORT_FORMATS:
        outputs["xlsx"] = export_excel(df, coverage, imp_rpt, meta_sheet)

    log.info("=" * 50)
    log.info("EXPORT COMPLETE")
    log.info("=" * 50)
    for fmt, path in outputs.items():
        log.info("  %-6s -> %s", fmt.upper(), path)

    return outputs


if __name__ == "__main__":
    run()
